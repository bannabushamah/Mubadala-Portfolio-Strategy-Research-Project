"""
build_excel_model.py
Creates Valuation_Models.xlsx - a LIVE workbook. The DCF sheets contain real
Excel formulas, not pasted values, so a reader can change a blue input cell
and watch the valuation move. That is the difference between a model and a
screenshot.

Colour convention (the standard used on every banking desk):
    BLUE  text = hard-coded input you may change
    BLACK text = formula calculated within the sheet
    GREEN text = link to another sheet
"""
import os, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

HERE = os.path.dirname(os.path.abspath(__file__))
OUT  = os.path.join(HERE, "Valuation_Models.xlsx")

NAVY  = "0B2545"; GOLD = "C9A227"; LIGHT = "EEF2F7"; WHITE = "FFFFFF"
BLUE_INPUT = Font(color="0000CC", name="Calibri", size=10)
BLACK_CALC = Font(color="000000", name="Calibri", size=10)
GREEN_LINK = Font(color="008000", name="Calibri", size=10)
H1 = Font(color=WHITE, bold=True, size=14, name="Calibri")
H2 = Font(color=NAVY, bold=True, size=11, name="Calibri")
FILL_H = PatternFill("solid", fgColor=NAVY)
FILL_S = PatternFill("solid", fgColor=LIGHT)
FILL_G = PatternFill("solid", fgColor=GOLD)
thin = Side(style="thin", color="BBBBBB")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook(); wb.remove(wb.active)

def sheet(title, header):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:J1")
    ws["A1"] = header; ws["A1"].font = H1; ws["A1"].fill = FILL_H
    ws["A1"].alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 26
    ws.column_dimensions["A"].width = 46
    for c in "BCDEFGHIJ": ws.column_dimensions[c].width = 14
    return ws

def lbl(ws, r, text, bold=False, indent=0):
    ws.cell(r,1,text).font = Font(bold=bold, size=10, name="Calibri",
                                  color=NAVY if bold else "000000")
    ws.cell(r,1).alignment = Alignment(indent=indent)

def put(ws, r, c, v, kind="calc", fmt=None):
    cell = ws.cell(r, c, v)
    cell.font = {"input":BLUE_INPUT,"calc":BLACK_CALC,"link":GREEN_LINK}[kind]
    if fmt: cell.number_format = fmt
    return cell

# =====================================================================  COVER
ws = sheet("Cover", "MUBADALA PORTFOLIO STRATEGY  |  VALUATION MODELS")
rows = [
 ("Project","Equity Research - Mubadala Portfolio Strategy Analyst"),
 ("Author","bann"),
 ("Valuation date","4 August 2026"),
 ("Currency","USD millions unless stated; OMV in EUR millions"),
 ("",""),
 ("CASE STUDIES",""),
 ("  1. Technology","GlobalFoundries Inc. (NASDAQ: GFS) - 5yr FCFF DCF + comparables + reverse DCF"),
 ("  2. Energy","OMV AG (VIE: OMV) - 5yr FCFF DCF + dividend cover + normalised P/E comps"),
 ("  3. Healthcare","WHOOP Inc. - reverse DCF on the disclosed Series G price"),
 ("  4. Financial Services","Mubadala Capital - fee-franchise sum-of-the-parts vs listed alt managers"),
 ("",""),
 ("HOW TO USE THIS WORKBOOK",""),
 ("  Blue figures","Inputs. Change them and every downstream number recalculates."),
 ("  Black figures","Formulas calculated on the same sheet."),
 ("  Green figures","Links to another sheet."),
 ("",""),
 ("HEALTH WARNING",""),
 ("  Forecasts","All forward assumptions are the author's, not company guidance."),
 ("  Private companies","WHOOP and Mubadala Capital outputs are frameworks built on labelled assumptions, not valuation opinions."),
 ("  Market data","Sourced from StockAnalysis / S&P Global Market Intelligence, frozen 3-4 August 2026."),
]
for i,(a,b) in enumerate(rows, start=3):
    bold = a.isupper() and a != ""
    lbl(ws, i, a, bold=bold)
    ws.cell(i,2,b).font = Font(size=10, name="Calibri"); ws.merge_cells(start_row=i,start_column=2,end_row=i,end_column=10)

# =====================================================================  GFS DCF
ws = sheet("GFS_DCF", "CASE 1  |  GLOBALFOUNDRIES INC. (NASDAQ: GFS)  -  FCFF DCF")
r = 3
lbl(ws, r, "MARKET DATA  (as at 3 Aug 2026)", bold=True); r += 1
market = [("Share price (USD)",50.01,"0.00"),("Shares outstanding (m)",548.70,"#,##0.0"),
          ("Market capitalisation (USDm)",27440.0,"#,##0"),("Enterprise value - market (USDm)",25400.0,"#,##0"),
          ("Gross debt (USDm)",1724.0,"#,##0"),("Revenue TTM (USDm)",6840.0,"#,##0"),
          ("EBITDA TTM (USDm)",2100.0,"#,##0"),("EBIT TTM (USDm)",826.0,"#,##0"),
          ("Equity beta (5y)",1.76,"0.00")]
first_market = r
for name,v,f in market:
    lbl(ws,r,"  "+name); put(ws,r,2,v,"input",f); r += 1
ws[f"B{first_market+9}"]  # placeholder
NET_CASH_ROW = r
lbl(ws,r,"  Net cash (mkt cap less EV) (USDm)"); put(ws,r,2,f"=B{first_market+2}-B{first_market+3}","calc","#,##0"); r += 2

lbl(ws, r, "COST OF CAPITAL", bold=True); r += 1
wacc_start = r
for name,v,f in [("Risk-free rate",0.0425,"0.00%"),("Equity risk premium",0.050,"0.00%"),
                 ("Beta used",1.76,"0.00"),("Cost of equity","=B{0}+B{1}*B{2}".format(r,r+1,r+2),"0.00%"),
                 ("Pre-tax cost of debt",0.055,"0.00%"),("Tax rate",0.15,"0.00%"),
                 ("After-tax cost of debt","=B{0}*(1-B{1})".format(r+4,r+5),"0.00%"),
                 ("Weight of equity","=B{0}/(B{0}+B{1})".format(first_market+2,first_market+4),"0.0%"),
                 ("Weight of debt","=1-B{0}".format(r+7),"0.0%"),
                 ("WACC","=B{0}*B{1}+B{2}*B{3}".format(r+7,r+3,r+8,r+6),"0.00%")]:
    lbl(ws,r,"  "+name)
    put(ws,r,2,v,"input" if isinstance(v,float) else "calc",f); r += 1
WACC_ROW = wacc_start + 9
TERM_G_ROW = r
lbl(ws,r,"  Terminal growth rate"); put(ws,r,2,0.025,"input","0.00%"); r += 2

lbl(ws, r, "FORECAST  (year 1 = FY2026E)", bold=True); r += 1
hdr = r
lbl(ws,r,"  USDm"); 
for i in range(5):
    c = ws.cell(r,2+i,f"Y{i+1}"); c.font=Font(bold=True,size=10,color=NAVY); c.fill=FILL_S
    c.alignment=Alignment(horizontal="center")
r += 1
GROWTH_ROW = r
lbl(ws,r,"  Revenue growth")
for i,v in enumerate([0.08,0.10,0.10,0.07,0.05]): put(ws,r,2+i,v,"input","0.0%")
r += 1
REV_ROW = r
lbl(ws,r,"  Revenue")
put(ws,r,2,f"=B{first_market+5}*(1+B{GROWTH_ROW})","calc","#,##0")
for i in range(1,5):
    col=get_column_letter(2+i); pcol=get_column_letter(1+i)
    put(ws,r,2+i,f"={pcol}{REV_ROW}*(1+{col}{GROWTH_ROW})","calc","#,##0")
r += 1
MARGIN_ROW = r
lbl(ws,r,"  EBIT margin")
for i,v in enumerate([0.130,0.150,0.170,0.190,0.200]): put(ws,r,2+i,v,"input","0.0%")
r += 1
EBIT_ROW = r
lbl(ws,r,"  EBIT")
for i in range(5):
    col=get_column_letter(2+i); put(ws,r,2+i,f"={col}{REV_ROW}*{col}{MARGIN_ROW}","calc","#,##0")
r += 1
NOPAT_ROW = r
lbl(ws,r,"  NOPAT  = EBIT x (1 - tax)")
for i in range(5):
    col=get_column_letter(2+i); put(ws,r,2+i,f"={col}{EBIT_ROW}*(1-$B${wacc_start+5})","calc","#,##0")
r += 1
DA_PCT_ROW = r
lbl(ws,r,"  D&A  (% of revenue)")
for i,v in enumerate([0.170,0.165,0.160,0.155,0.150]): put(ws,r,2+i,v,"input","0.0%")
r += 1
DA_ROW = r
lbl(ws,r,"  Depreciation & amortisation")
for i in range(5):
    col=get_column_letter(2+i); put(ws,r,2+i,f"={col}{REV_ROW}*{col}{DA_PCT_ROW}","calc","#,##0")
r += 1
CX_PCT_ROW = r
lbl(ws,r,"  Capex  (% of revenue)")
for i,v in enumerate([0.130,0.135,0.135,0.130,0.125]): put(ws,r,2+i,v,"input","0.0%")
r += 1
CX_ROW = r
lbl(ws,r,"  Capital expenditure")
for i in range(5):
    col=get_column_letter(2+i); put(ws,r,2+i,f"=-{col}{REV_ROW}*{col}{CX_PCT_ROW}","calc","#,##0")
r += 1
NWC_ROW = r
lbl(ws,r,"  Change in net working capital")
put(ws,r,2,f"=-(B{REV_ROW}-B{first_market+5})*$B${NWC_ROW+1}","calc","#,##0")
for i in range(1,5):
    col=get_column_letter(2+i); pcol=get_column_letter(1+i)
    put(ws,r,2+i,f"=-({col}{REV_ROW}-{pcol}{REV_ROW})*$B${NWC_ROW+1}","calc","#,##0")
r += 1
lbl(ws,r,"  NWC investment per $1 of new revenue"); put(ws,r,2,0.02,"input","0.0%"); r += 1
FCFF_ROW = r
lbl(ws,r,"  FREE CASH FLOW TO FIRM", bold=True)
for i in range(5):
    col=get_column_letter(2+i)
    c=put(ws,r,2+i,f"={col}{NOPAT_ROW}+{col}{DA_ROW}+{col}{CX_ROW}+{col}{NWC_ROW}","calc","#,##0")
    c.font=Font(bold=True,size=10); c.fill=FILL_S
r += 1
DF_ROW = r
lbl(ws,r,"  Discount factor (mid-year)")
for i in range(5):
    put(ws,r,2+i,f"=1/(1+$B${WACC_ROW})^({i+1}-0.5)","calc","0.000")
r += 1
PV_ROW = r
lbl(ws,r,"  PV of FCFF")
for i in range(5):
    col=get_column_letter(2+i); put(ws,r,2+i,f"={col}{FCFF_ROW}*{col}{DF_ROW}","calc","#,##0")
r += 2

lbl(ws, r, "VALUATION BRIDGE", bold=True); r += 1
VB = r
items = [
 ("Sum of PV of explicit FCFF", f"=SUM(B{PV_ROW}:F{PV_ROW})"),
 ("Terminal value (Gordon growth)", f"=F{FCFF_ROW}*(1+B{TERM_G_ROW})/(B{WACC_ROW}-B{TERM_G_ROW})"),
 ("PV of terminal value", f"=B{VB+1}/(1+B{WACC_ROW})^4.5"),
 ("ENTERPRISE VALUE", f"=B{VB}+B{VB+2}"),
 ("Terminal value as % of EV", f"=B{VB+2}/B{VB+3}"),
 ("Plus: net cash", f"=B{NET_CASH_ROW}"),
 ("EQUITY VALUE", f"=B{VB+3}+B{VB+5}"),
 ("Shares outstanding (m)", f"=B{first_market+1}"),
 ("VALUE PER SHARE (USD)", f"=B{VB+6}/B{VB+7}"),
 ("Current share price (USD)", f"=B{first_market}"),
 ("Upside / (downside)", f"=B{VB+8}/B{VB+9}-1"),
]
for i,(n,f) in enumerate(items):
    bold = n.isupper()
    lbl(ws, VB+i, "  "+n, bold=bold)
    fmt = "0.0%" if "%" in n or "Upside" in n else ("0.00" if "SHARE" in n or "price" in n else "#,##0")
    c = put(ws, VB+i, 2, f, "calc", fmt)
    if bold: c.font=Font(bold=True,size=10,color=NAVY); c.fill=FILL_G
r = VB + len(items) + 2

lbl(ws, r, "CROSS-CHECK: TRADING COMPARABLES", bold=True); r += 1
CC = r
lbl(ws,r,"  Applied EV/EBITDA - low");    put(ws,r,2,11.0,"input","0.0x"); r+=1
lbl(ws,r,"  Applied EV/EBITDA - mid");    put(ws,r,2,14.5,"input","0.0x"); r+=1
lbl(ws,r,"  Applied EV/EBITDA - high");   put(ws,r,2,17.0,"input","0.0x"); r+=1
lbl(ws,r,"  Implied value/share - low");  put(ws,r,2,f"=(B{CC}*B{first_market+6}+B{NET_CASH_ROW})/B{first_market+1}","calc","0.00"); r+=1
lbl(ws,r,"  Implied value/share - mid");  put(ws,r,2,f"=(B{CC+1}*B{first_market+6}+B{NET_CASH_ROW})/B{first_market+1}","calc","0.00"); r+=1
lbl(ws,r,"  Implied value/share - high"); put(ws,r,2,f"=(B{CC+2}*B{first_market+6}+B{NET_CASH_ROW})/B{first_market+1}","calc","0.00"); r+=1
lbl(ws,r,"  Note"); ws.cell(r,2,"Tower Semiconductor (90.8x) excluded as a momentum outlier").font=Font(italic=True,size=9)

# =====================================================================  OMV DCF
ws = sheet("OMV_DCF", "CASE 2  |  OMV AG (VIE: OMV)  -  FCFF DCF  (EUR millions)")
r = 3
lbl(ws,r,"MARKET DATA  (as at 30 Jun 2026)",bold=True); r+=1
m0 = r
for name,v,f in [("Share price (EUR)",54.60,"0.00"),("Shares outstanding (m)",326.0,"#,##0.0"),
                 ("Market capitalisation (EURm)",17620.0,"#,##0"),("Enterprise value - market (EURm)",24160.0,"#,##0"),
                 ("Gross debt (EURm)",8010.0,"#,##0"),("Cash (EURm)",5220.0,"#,##0"),
                 ("Revenue TTM (EURm)",23378.0,"#,##0"),("EBITDA TTM (EURm)",5402.0,"#,##0"),
                 ("Dividend per share (EUR)",4.40,"0.00"),("Raw 5-year beta (rejected)",0.21,"0.00")]:
    lbl(ws,r,"  "+name); put(ws,r,2,v,"input",f); r+=1
ND_ROW = r; lbl(ws,r,"  Net debt (EURm)"); put(ws,r,2,f"=B{m0+4}-B{m0+5}","calc","#,##0"); r+=1
MI_ROW = r; lbl(ws,r,"  Minority interests (implied by EV bridge)")
put(ws,r,2,f"=B{m0+3}-B{m0+2}-B{ND_ROW}","calc","#,##0"); r+=2

lbl(ws,r,"COST OF CAPITAL",bold=True); r+=1
w0=r
for name,v,f in [("Risk-free rate (10y Bund)",0.0275,"0.00%"),("Equity risk premium",0.055,"0.00%"),
                 ("Sector beta used (not raw beta)",0.90,"0.00"),("Country / commodity premium",0.010,"0.00%"),
                 ("Cost of equity",f"=B{r}+B{r+1}*B{r+2}+B{r+3}","0.00%"),
                 ("Pre-tax cost of debt",0.045,"0.00%"),("Tax rate (normalised)",0.30,"0.00%"),
                 ("After-tax cost of debt",f"=B{r+5}*(1-B{r+6})","0.00%"),
                 ("Weight of equity",f"=B{m0+2}/(B{m0+2}+B{m0+4})","0.0%"),
                 ("Weight of debt",f"=1-B{r+8}","0.0%"),
                 ("WACC",f"=B{r+8}*B{r+4}+B{r+9}*B{r+7}","0.00%")]:
    lbl(ws,r,"  "+name); put(ws,r,2,v,"input" if isinstance(v,float) else "calc",f); r+=1
OW=w0+10
OG=r; lbl(ws,r,"  Terminal growth rate"); put(ws,r,2,0.010,"input","0.00%"); r+=2

lbl(ws,r,"FORECAST",bold=True); r+=1
for i in range(5):
    c=ws.cell(r,2+i,f"Y{i+1}"); c.font=Font(bold=True,size=10,color=NAVY); c.fill=FILL_S
    c.alignment=Alignment(horizontal="center")
lbl(ws,r,"  EURm"); r+=1
OGR=r; lbl(ws,r,"  Revenue growth")
for i,v in enumerate([-0.03,0.01,0.02,0.02,0.02]): put(ws,r,2+i,v,"input","0.0%")
r+=1
ORV=r; lbl(ws,r,"  Revenue")
put(ws,r,2,f"=B{m0+6}*(1+B{OGR})","calc","#,##0")
for i in range(1,5):
    col=get_column_letter(2+i); p=get_column_letter(1+i)
    put(ws,r,2+i,f"={p}{ORV}*(1+{col}{OGR})","calc","#,##0")
r+=1
OMG=r; lbl(ws,r,"  EBIT margin")
for i,v in enumerate([0.140,0.145,0.150,0.150,0.150]): put(ws,r,2+i,v,"input","0.0%")
r+=1
OEB=r; lbl(ws,r,"  EBIT")
for i in range(5):
    c=get_column_letter(2+i); put(ws,r,2+i,f"={c}{ORV}*{c}{OMG}","calc","#,##0")
r+=1
ONO=r; lbl(ws,r,"  NOPAT")
for i in range(5):
    c=get_column_letter(2+i); put(ws,r,2+i,f"={c}{OEB}*(1-$B${w0+6})","calc","#,##0")
r+=1
ODP=r; lbl(ws,r,"  D&A (% of revenue)")
for i,v in enumerate([0.080,0.079,0.078,0.077,0.076]): put(ws,r,2+i,v,"input","0.0%")
r+=1
ODA=r; lbl(ws,r,"  Depreciation & amortisation")
for i in range(5):
    c=get_column_letter(2+i); put(ws,r,2+i,f"={c}{ORV}*{c}{ODP}","calc","#,##0")
r+=1
OCP=r; lbl(ws,r,"  Capex (% of revenue)")
for i,v in enumerate([0.150,0.135,0.125,0.118,0.115]): put(ws,r,2+i,v,"input","0.0%")
r+=1
OCX=r; lbl(ws,r,"  Capital expenditure")
for i in range(5):
    c=get_column_letter(2+i); put(ws,r,2+i,f"=-{c}{ORV}*{c}{OCP}","calc","#,##0")
r+=1
ONW=r; lbl(ws,r,"  Change in net working capital")
put(ws,r,2,f"=-(B{ORV}-B{m0+6})*$B${ONW+1}","calc","#,##0")
for i in range(1,5):
    c=get_column_letter(2+i); p=get_column_letter(1+i)
    put(ws,r,2+i,f"=-({c}{ORV}-{p}{ORV})*$B${ONW+1}","calc","#,##0")
r+=1
lbl(ws,r,"  NWC investment per EUR1 of new revenue"); put(ws,r,2,0.03,"input","0.0%"); r+=1
OFC=r; lbl(ws,r,"  FREE CASH FLOW TO FIRM",bold=True)
for i in range(5):
    c=get_column_letter(2+i)
    cell=put(ws,r,2+i,f"={c}{ONO}+{c}{ODA}+{c}{OCX}+{c}{ONW}","calc","#,##0")
    cell.font=Font(bold=True,size=10); cell.fill=FILL_S
r+=1
ODF=r; lbl(ws,r,"  Discount factor (mid-year)")
for i in range(5): put(ws,r,2+i,f"=1/(1+$B${OW})^({i+1}-0.5)","calc","0.000")
r+=1
OPV=r; lbl(ws,r,"  PV of FCFF")
for i in range(5):
    c=get_column_letter(2+i); put(ws,r,2+i,f"={c}{OFC}*{c}{ODF}","calc","#,##0")
r+=2
lbl(ws,r,"VALUATION BRIDGE",bold=True); r+=1
V=r
for i,(n,f) in enumerate([
  ("Sum of PV of explicit FCFF", f"=SUM(B{OPV}:F{OPV})"),
  ("Terminal value", f"=F{OFC}*(1+B{OG})/(B{OW}-B{OG})"),
  ("PV of terminal value", f"=B{V+1}/(1+B{OW})^4.5"),
  ("ENTERPRISE VALUE", f"=B{V}+B{V+2}"),
  ("Less: net debt", f"=-B{ND_ROW}"),
  ("Less: minority interests", f"=-B{MI_ROW}"),
  ("EQUITY VALUE", f"=B{V+3}+B{V+4}+B{V+5}"),
  ("VALUE PER SHARE (EUR)", f"=B{V+6}/B{m0+1}"),
  ("Current share price (EUR)", f"=B{m0}"),
  ("Upside / (downside)", f"=B{V+7}/B{V+8}-1"),
  ("Dividend yield at current price", f"=B{m0+8}/B{m0}")]):
    bold=n.isupper(); lbl(ws,V+i,"  "+n,bold=bold)
    fmt="0.0%" if ("Upside" in n or "yield" in n) else ("0.00" if "SHARE" in n or "price" in n else "#,##0")
    c=put(ws,V+i,2,f,"calc",fmt)
    if bold: c.font=Font(bold=True,size=10,color=NAVY); c.fill=FILL_G

# ================================================================  DIVIDEND TEST
ws = sheet("OMV_Dividend_Test", "CASE 2  |  OMV  -  IS THE 8.1% DIVIDEND SAFE?")
hdr=["Year","Free cash flow (EURm)","DPS (EUR)","Shares (m)","Dividend cost (EURm)","FCF cover (x)","Verdict"]
for j,h in enumerate(hdr,1):
    c=ws.cell(3,j,h); c.font=Font(bold=True,color=WHITE,size=10); c.fill=FILL_H
data=[(2021,4520,2.30),(2022,4815,2.80),(2023,2222,2.95),(2024,1943,3.05),(2025,1366,3.15)]
for i,(y,f,d) in enumerate(data):
    rr=4+i
    put(ws,rr,1,y,"input","0"); put(ws,rr,2,f,"input","#,##0"); put(ws,rr,3,d,"input","0.00")
    put(ws,rr,4,326.0,"input","#,##0")
    put(ws,rr,5,f"=C{rr}*D{rr}","calc","#,##0")
    put(ws,rr,6,f"=B{rr}/E{rr}","calc","0.00")
    put(ws,rr,7,f'=IF(F{rr}<1,"NOT COVERED",IF(F{rr}<1.5,"THIN","COMFORTABLE"))',"calc")
ws["A10"]="Finding"; ws["A10"].font=H2
ws.merge_cells("B10:G12")
ws["B10"]=("Free-cash-flow cover of the ordinary dividend has fallen from 6.0x in 2021 to 1.3x in 2025, "
           "while the dividend per share has been increased every single year. The payout is still covered, "
           "but the margin of safety is largely gone. One further leg down in refining or chemicals margins "
           "and the distribution has to be funded from the balance sheet.")
ws["B10"].alignment=Alignment(wrap_text=True, vertical="top")
ch=LineChart(); ch.title="OMV: FCF cover of the dividend (x)"; ch.height=7; ch.width=16
ch.add_data(Reference(ws,min_col=6,min_row=3,max_row=8),titles_from_data=True)
ch.set_categories(Reference(ws,min_col=1,min_row=4,max_row=8))
ws.add_chart(ch,"A14")

# ================================================================  WHOOP
ws = sheet("WHOOP_Reverse_DCF", "CASE 3  |  WHOOP INC.  -  WHAT DOES THE SERIES G PRICE ASSUME?")
r=3
lbl(ws,r,"DISCLOSED FACTS  (Mubadala press release, 31 Mar 2026)",bold=True); r+=1
w=r
for n,v,f in [("Post-money valuation (USDm)",10100,"#,##0"),("Round size (USDm)",575,"#,##0"),
              ("Exit-2025 bookings run-rate (USDm)",1100,"#,##0"),("Members (m)",2.5,"0.0"),
              ("2025 bookings growth",1.03,"0.0%")]:
    lbl(ws,r,"  "+n); put(ws,r,2,v,"input",f); r+=1
lbl(ws,r,"  Pre-money valuation (USDm)"); put(ws,r,2,f"=B{w}-B{w+1}","calc","#,##0"); PRE=r; r+=1
lbl(ws,r,"  Implied EV / revenue - post-money"); put(ws,r,2,f"=B{w}/B{w+2}","calc","0.0x"); r+=1
lbl(ws,r,"  Implied EV / revenue - pre-money"); put(ws,r,2,f"=B{PRE}/B{w+2}","calc","0.0x"); r+=1
lbl(ws,r,"  Revenue per member (USD)"); put(ws,r,2,f"=B{w+2}/B{w+3}","calc","#,##0"); r+=2
lbl(ws,r,"ASSUMPTIONS  (the author's, not the company's)",bold=True); r+=1
a=r
for n,v,f in [("Required return (late-stage private)",0.18,"0.0%"),("Terminal growth",0.035,"0.0%"),
              ("Horizon (years)",10,"0"),("Starting FCF margin",0.04,"0.0%"),
              ("Terminal FCF margin",0.20,"0.0%"),("Revenue CAGR implied by the price",0.348,"0.0%")]:
    lbl(ws,r,"  "+n); put(ws,r,2,v,"input",f); r+=1
r+=1
lbl(ws,r,"IMPLIED REVENUE PATH",bold=True); r+=1
hr=r; ws.cell(r,1,"  Year").font=Font(bold=True,size=10,color=NAVY)
for j,h in enumerate(["Revenue (USDm)","FCF margin","FCF (USDm)","Discount factor","PV (USDm)"],2):
    c=ws.cell(r,j,h); c.font=Font(bold=True,size=10,color=NAVY); c.fill=FILL_S
r+=1
start=r
for t in range(1,11):
    rr=start+t-1
    put(ws,rr,1,t,"calc","0")
    prev = f"B{rr-1}" if t>1 else f"B{w+2}"
    put(ws,rr,2,f"={prev}*(1+$B${a+5})","calc","#,##0")
    put(ws,rr,3,f"=$B${a+3}+($B${a+4}-$B${a+3})*(A{rr}/$B${a+2})","calc","0.0%")
    put(ws,rr,4,f"=B{rr}*C{rr}","calc","#,##0")
    put(ws,rr,5,f"=1/(1+$B${a})^A{rr}","calc","0.000")
    put(ws,rr,6,f"=D{rr}*E{rr}","calc","#,##0")
r=start+11
lbl(ws,r,"  Sum of PV of explicit FCF"); put(ws,r,2,f"=SUM(F{start}:F{start+9})","calc","#,##0"); r+=1
lbl(ws,r,"  Terminal value"); put(ws,r,2,f"=B{start+9}*$B${a+4}*(1+$B${a+1})/($B${a}-$B${a+1})","calc","#,##0"); TV=r; r+=1
lbl(ws,r,"  PV of terminal value"); put(ws,r,2,f"=B{TV}/(1+$B${a})^10","calc","#,##0"); r+=1
lbl(ws,r,"  IMPLIED VALUE (USDm)",bold=True)
c=put(ws,r,2,f"=B{r-3}+B{r-1}","calc","#,##0"); c.font=Font(bold=True,size=10,color=NAVY); c.fill=FILL_G; r+=1
lbl(ws,r,"  Price actually paid, pre-money (USDm)"); put(ws,r,2,f"=B{PRE}","calc","#,##0"); r+=1
lbl(ws,r,"  Difference"); put(ws,r,2,f"=B{r-2}/B{r-1}-1","calc","0.0%"); r+=2
ws.cell(r,1,"Reading").font=H2
ws.merge_cells(start_row=r,start_column=2,end_row=r+2,end_column=8)
ws.cell(r,2,"The Series G price only makes sense if revenue compounds at roughly 35% a year for a decade - "
            "taking WHOOP from a $1.1bn run-rate to about $22bn - while free-cash-flow margins climb to 20%. "
            "That is a demanding but not impossible path given 2025 bookings growth of 103%. The investment "
            "question is therefore about the DURABILITY of growth, not its direction.").alignment=Alignment(wrap_text=True,vertical="top")

# ================================================================  MUBADALA CAPITAL
ws = sheet("Mubadala_Capital", "CASE 4  |  MUBADALA CAPITAL  -  WHAT IS THE FEE FRANCHISE WORTH?")
r=3
lbl(ws,r,"THE ONLY HARD FACT",bold=True); r+=1
lbl(ws,r,"  Aggregate AUM (USDbn) - mubadala.com"); put(ws,r,2,30.0,"input","0.0"); AUM=r; r+=2
lbl(ws,r,"ASSUMPTIONS",bold=True); r+=1
A0=r
for n,v,f in [("Third-party share of AUM",0.60,"0.0%"),("Blended management fee rate",0.0120,"0.00%"),
              ("Fee-related-earnings margin",0.35,"0.0%"),("Private-company discount",0.25,"0.0%")]:
    lbl(ws,r,"  "+n); put(ws,r,2,v,"input",f); r+=1
r+=1
lbl(ws,r,"DERIVED",bold=True); r+=1
D0=r
lbl(ws,r,"  Third-party AUM (USDbn)"); put(ws,r,2,f"=B{AUM}*B{A0}","calc","0.0"); r+=1
lbl(ws,r,"  Management fee revenue (USDm)"); put(ws,r,2,f"=B{D0}*1000*B{A0+1}","calc","#,##0"); FEE=r; r+=1
lbl(ws,r,"  Fee-related earnings (USDm)"); put(ws,r,2,f"=B{FEE}*B{A0+2}","calc","#,##0"); FRE=r; r+=2
lbl(ws,r,"LISTED ALTERNATIVE-MANAGER COHORT",bold=True); r+=1
for j,h in enumerate(["Manager","Market cap (USDbn)","Revenue (USDbn)","Price / sales","P/E"],1):
    c=ws.cell(r,j,h); c.font=Font(bold=True,color=WHITE,size=10); c.fill=FILL_H
r+=1
cohort=[("Blackstone (BX)",169.49,15.48,30.51),("Apollo (APO) - excluded",75.06,35.60,81.92),
        ("Ares (ARES)",46.56,5.99,62.43),("Blue Owl (OWL)",17.80,2.99,95.88),("TPG (TPG)",18.74,3.73,139.03)]
c0=r
for i,(n,mc,rev,pe) in enumerate(cohort):
    rr=r+i
    ws.cell(rr,1,n).font=Font(size=10,italic="excluded" in n)
    put(ws,rr,2,mc,"input","#,##0.0"); put(ws,rr,3,rev,"input","#,##0.00")
    put(ws,rr,4,f"=B{rr}/C{rr}","calc","0.00"); put(ws,rr,5,pe,"input","0.0")
r=c0+len(cohort)+1
lbl(ws,r,"  Median price/sales (excluding Apollo)")
put(ws,r,2,f"=MEDIAN(D{c0},D{c0+2},D{c0+3},D{c0+4})","calc","0.00"); PSM=r; r+=1
lbl(ws,r,"  Median P/E"); put(ws,r,2,f"=MEDIAN(E{c0}:E{c0+4})","calc","0.0"); PEM=r; r+=2
lbl(ws,r,"VALUATION RANGE (USDbn)",bold=True); r+=1
for n,f in [("Low  - 5.0x fee revenue", f"=B{FEE}*5/1000"),
            ("Mid  - peer median price/sales", f"=B{FEE}*B{PSM}/1000"),
            ("High - 10.9x fee revenue (Blackstone)", f"=B{FEE}*10.9/1000"),
            ("Cross-check - 30.5x FRE (Blackstone P/E)", f"=B{FRE}*30.51/1000")]:
    lbl(ws,r,"  "+n); put(ws,r,2,f,"calc","0.00"); r+=1
lbl(ws,r,"  CENTRAL VALUE before discount", bold=True)
c=put(ws,r,2,f"=MEDIAN(B{r-4}:B{r-1})","calc","0.00"); r+=1
lbl(ws,r,"  CENTRAL VALUE after private-company discount", bold=True)
c=put(ws,r,2,f"=B{r-1}*(1-B{A0+3})","calc","0.00"); c.font=Font(bold=True,size=10,color=NAVY); c.fill=FILL_G

# ================================================================  SOURCES
ws = sheet("Source_Register", "SOURCE REGISTER  -  EVERY NUMBER IN THIS WORKBOOK IS TRACEABLE")
for j,h in enumerate(["ID","Source","URL","As-of date","Class"],1):
    c=ws.cell(3,j,h); c.font=Font(bold=True,color=WHITE,size=10); c.fill=FILL_H
ws.column_dimensions["B"].width=52; ws.column_dimensions["C"].width=70
ws.column_dimensions["D"].width=14; ws.column_dimensions["E"].width=18
import csv as _csv
src=os.path.join(os.path.dirname(HERE),"01_data","warehouse","dim_source.csv")
with open(src) as f:
    for i,row in enumerate(list(_csv.reader(f))[1:]):
        for j,v in enumerate(row,1): ws.cell(4+i,j,v).font=Font(size=9)

wb.save(OUT)
print("Written:", OUT)
